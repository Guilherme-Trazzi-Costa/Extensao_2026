
import csv
import json
import math
import re
import sys
from dataclasses import dataclass

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None
    filedialog = None
    messagebox = None
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, StyleSheet1, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ============================================================
# Configuração padrão
# ============================================================

DEFAULT_CONFIG: Dict[str, Any] = {
    "branding": {
        "report_title": "Relatório Financeiro Executivo",
        "subtitle": "Síntese gerada a partir da saída do Insight Calculado",
        "company": "Insight Calculado",
        "accent_color": "#0F4C81",
        "secondary_color": "#EAF2F8",
        "text_color": "#1F2937",
        "muted_color": "#6B7280",
        "confidentiality": "Uso interno",
        "footer_text": "Relatório gerado automaticamente pelo módulo de relatórios do Insight Calculado",
    },
    "pdf": {
        "include_cover": True,
        "include_charts": True,
        "include_lesson_breakdown": True,
        "include_methodology": True,
        "currency_symbol": "R$",
        "show_confidentiality_badge": True,
    },
    "excel": {
        "freeze_header": True,
        "sheet_color": "0F4C81",
    },
    "csv": {
        "delimiter": ";",
    },
}


# ============================================================
# Utilidades
# ============================================================

def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: Dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    result = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9áàâãéèêíïóôõöúçñ\- ]", "", text, flags=re.IGNORECASE)
    text = text.replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text.strip("_") or "relatorio"


def fmt_number(value: Optional[float], digits: int = 2) -> str:
    if value is None:
        return "-"
    return f"{value:,.{digits}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_currency(value: Optional[float], symbol: str = "R$") -> str:
    if value is None:
        return "-"
    return f"{symbol} {fmt_number(value, 2)}"


def fmt_percent(value: Optional[float], digits: int = 2) -> str:
    if value is None:
        return "-"
    return f"{fmt_number(value * 100, digits)}%"


def safe_get(d: Dict[str, Any], *path: str, default=None):
    current = d
    for key in path:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
        if current is None:
            return default
    return current


def flatten_dict(data: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
    flat = {}
    for k, v in data.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            flat.update(flatten_dict(v, key))
        else:
            flat[key] = v
    return flat


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def hex_color(value: str):
    return HexColor(value)


def pick(series: List[Any], idx: int, default=None):
    return series[idx] if idx < len(series) else default


# ============================================================
# Modelo de dados
# ============================================================

@dataclass
class ReportContext:
    raw: Dict[str, Any]
    config: Dict[str, Any]
    negocio_nome: str
    segmento: str
    cidade: str
    kpis: Dict[str, Any]
    alertas: List[str]
    top_categorias: List[Dict[str, Any]]
    series: Dict[str, List[float]]
    correlacoes: Dict[str, Any]
    aula_cards: List[Dict[str, Any]]
    created_at: str


def build_context(data: Dict[str, Any], config: Dict[str, Any]) -> ReportContext:
    negocio = data.get("negocio", {})
    modulo = safe_get(data, "modulo_final", "saida_principal", default={}) or {}
    aula_4_corr = safe_get(data, "aula_4", "calculos", "correlacoes", default={}) or {}
    aula_cards = []
    for i in range(1, 9):
        aula = data.get(f"aula_{i}", {})
        if not aula:
            continue
        insights = aula.get("insights", [])
        aula_cards.append({
            "aula": f"Aula {i}",
            "tema": aula.get("tema", ""),
            "problema": aula.get("problema_financeiro", ""),
            "insight_principal": insights[0] if insights else "",
        })

    return ReportContext(
        raw=data,
        config=config,
        negocio_nome=negocio.get("nome", "Negócio não informado"),
        segmento=negocio.get("segmento", "Segmento não informado"),
        cidade=negocio.get("cidade", "Cidade não informada"),
        kpis=modulo.get("kpis_finais", {}),
        alertas=modulo.get("alertas", []),
        top_categorias=modulo.get("top_categorias", []),
        series=modulo.get("dados_prontos_para_graficos", {}),
        correlacoes=aula_4_corr,
        aula_cards=aula_cards,
        created_at=datetime.now().strftime("%d/%m/%Y %H:%M"),
    )


# ============================================================
# Preparação dos datasets tabulares
# ============================================================

def build_summary_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    return [
        {"metrica": "Negócio", "valor": ctx.negocio_nome},
        {"metrica": "Segmento", "valor": ctx.segmento},
        {"metrica": "Cidade", "valor": ctx.cidade},
        {"metrica": "Receita total", "valor": ctx.kpis.get("receita_total")},
        {"metrica": "Despesa total", "valor": ctx.kpis.get("despesa_total")},
        {"metrica": "Lucro total", "valor": ctx.kpis.get("lucro_total")},
        {"metrica": "Margem de lucro", "valor": ctx.kpis.get("margem_lucro")},
        {"metrica": "Ticket médio", "valor": ctx.kpis.get("ticket_medio")},
        {"metrica": "Probabilidade de atraso", "valor": ctx.kpis.get("probabilidade_atraso_pagamento")},
        {"metrica": "Probabilidade de lucro negativo", "valor": ctx.kpis.get("probabilidade_lucro_negativo")},
        {"metrica": "Data de geração", "valor": ctx.created_at},
    ]


def build_daily_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    receita = ctx.series.get("receita_diaria", [])
    despesa = ctx.series.get("despesa_diaria", [])
    lucro = ctx.series.get("lucro_diario", [])
    vendas = ctx.series.get("vendas_qtd_diaria", [])
    n = max(len(receita), len(despesa), len(lucro), len(vendas))
    rows = []
    for i in range(n):
        rows.append({
            "periodo": f"Dia {i+1}",
            "receita": pick(receita, i),
            "despesa": pick(despesa, i),
            "lucro": pick(lucro, i),
            "vendas_qtd": pick(vendas, i),
        })
    return rows


def build_kpi_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    return [
        {"indicador": "Receita total", "valor": ctx.kpis.get("receita_total"), "formato": "currency"},
        {"indicador": "Despesa total", "valor": ctx.kpis.get("despesa_total"), "formato": "currency"},
        {"indicador": "Lucro total", "valor": ctx.kpis.get("lucro_total"), "formato": "currency"},
        {"indicador": "Margem de lucro", "valor": ctx.kpis.get("margem_lucro"), "formato": "percent"},
        {"indicador": "Ticket médio", "valor": ctx.kpis.get("ticket_medio"), "formato": "currency"},
        {"indicador": "Probabilidade de atraso", "valor": ctx.kpis.get("probabilidade_atraso_pagamento"), "formato": "percent"},
        {"indicador": "Probabilidade de lucro negativo", "valor": ctx.kpis.get("probabilidade_lucro_negativo"), "formato": "percent"},
    ]


def build_correlation_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    rows = []
    for nome, valor in ctx.correlacoes.items():
        rows.append({
            "relacao": nome,
            "correlacao": valor,
            "forca": interpret_correlation(valor),
        })
    return rows


def interpret_correlation(value: Optional[float]) -> str:
    if value is None:
        return "Indisponível"
    strength = abs(value)
    if strength >= 0.8:
        return "Muito forte"
    if strength >= 0.6:
        return "Forte"
    if strength >= 0.4:
        return "Moderada"
    if strength >= 0.2:
        return "Fraca"
    return "Muito fraca"


def build_top_categories_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    return [
        {
            "categoria": item.get("valor"),
            "frequencia": item.get("frequencia"),
            "percentual": item.get("percentual"),
        }
        for item in ctx.top_categorias
    ]


def build_lessons_rows(ctx: ReportContext) -> List[Dict[str, Any]]:
    return ctx.aula_cards


# ============================================================
# Exportação CSV
# ============================================================

def write_csv(path: Path, rows: List[Dict[str, Any]], delimiter: str = ";") -> None:
    if not rows:
        rows = [{"info": "sem dados"}]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def export_csv_bundle(ctx: ReportContext, output_dir: Path, prefix: str) -> List[Path]:
    ensure_dir(output_dir)
    delimiter = ctx.config["csv"]["delimiter"]
    files = []
    datasets = {
        f"{prefix}_resumo.csv": build_summary_rows(ctx),
        f"{prefix}_series_diarias.csv": build_daily_rows(ctx),
        f"{prefix}_kpis.csv": build_kpi_rows(ctx),
        f"{prefix}_correlacoes.csv": build_correlation_rows(ctx),
        f"{prefix}_top_categorias.csv": build_top_categories_rows(ctx),
        f"{prefix}_aulas.csv": build_lessons_rows(ctx),
    }
    for name, rows in datasets.items():
        path = output_dir / name
        write_csv(path, rows, delimiter=delimiter)
        files.append(path)
    return files


# ============================================================
# Exportação Excel
# ============================================================

def style_header(ws, row=1, fill_color="0F4C81"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws, extra=2):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + extra, 42)


def add_sheet_from_rows(wb: Workbook, title: str, rows: List[Dict[str, Any]], fill_color="0F4C81"):
    ws = wb.create_sheet(title=title)
    if not rows:
        rows = [{"info": "sem dados"}]
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    style_header(ws, 1, fill_color)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    auto_fit_columns(ws)
    ws.freeze_panes = "A2"
    return ws


def export_excel(ctx: ReportContext, path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    fill_color = ctx.config["excel"]["sheet_color"]

    # Resumo executivo
    ws = wb.create_sheet("Resumo Executivo")
    ws["A1"] = ctx.config["branding"]["report_title"]
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=fill_color)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")

    meta_rows = [
        ("Negócio", ctx.negocio_nome),
        ("Segmento", ctx.segmento),
        ("Cidade", ctx.cidade),
        ("Gerado em", ctx.created_at),
    ]
    start = 3
    for i, (label, value) in enumerate(meta_rows, start=start):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value

    kpis = build_kpi_rows(ctx)
    row = start + len(meta_rows) + 2
    ws[f"A{row}"] = "KPIs"
    ws[f"A{row}"].font = Font(size=13, bold=True, color=fill_color)
    row += 1
    ws.append(["Indicador", "Valor", "Formato"])
    for item in kpis:
        raw_val = item["valor"]
        if item["formato"] == "currency":
            show_val = fmt_currency(raw_val)
        elif item["formato"] == "percent":
            show_val = fmt_percent(raw_val)
        else:
            show_val = raw_val
        ws.append([item["indicador"], show_val, item["formato"]])

    style_header(ws, row, fill_color)
    auto_fit_columns(ws)

    # Outras abas
    add_sheet_from_rows(wb, "Series Diarias", build_daily_rows(ctx), fill_color)
    add_sheet_from_rows(wb, "Correlacoes", build_correlation_rows(ctx), fill_color)
    add_sheet_from_rows(wb, "Top Categorias", build_top_categories_rows(ctx), fill_color)
    add_sheet_from_rows(wb, "Aulas", build_lessons_rows(ctx), fill_color)
    add_sheet_from_rows(wb, "Resumo Flat", build_summary_rows(ctx), fill_color)

    # Pequenos ajustes de número
    for ws_name in ["Series Diarias", "Correlacoes", "Top Categorias", "Resumo Flat"]:
        ws2 = wb[ws_name]
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)
    return path


# ============================================================
# Geração dos gráficos
# ============================================================

def create_line_chart(ctx: ReportContext, path: Path) -> Path:
    receita = ctx.series.get("receita_diaria", [])
    despesa = ctx.series.get("despesa_diaria", [])
    lucro = ctx.series.get("lucro_diario", [])
    x = list(range(1, max(len(receita), len(despesa), len(lucro)) + 1))

    plt.figure(figsize=(10, 4.5))
    if receita:
        plt.plot(x[:len(receita)], receita, marker="o", linewidth=2, label="Receita")
    if despesa:
        plt.plot(x[:len(despesa)], despesa, marker="o", linewidth=2, label="Despesa")
    if lucro:
        plt.plot(x[:len(lucro)], lucro, marker="o", linewidth=2, label="Lucro")
    plt.title("Séries financeiras por período")
    plt.xlabel("Período")
    plt.ylabel("Valor")
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()
    plt.savefig(path, dpi=180, bbox_inches="tight")
    plt.close()
    return path


def create_top_categories_chart(ctx: ReportContext, path: Path) -> Path:
    cats = build_top_categories_rows(ctx)
    labels = [r["categoria"] for r in cats[:6]]
    values = [r["frequencia"] for r in cats[:6]]

    plt.figure(figsize=(9, 4.5))
    plt.bar(labels, values)
    plt.title("Top categorias observadas")
    plt.xlabel("Categoria")
    plt.ylabel("Frequência")
    plt.xticks(rotation=20, ha="right")
    plt.tight_layout()
    plt.savefig(path, dpi=180, bbox_inches="tight")
    plt.close()
    return path


# ============================================================
# Exportação PDF
# ============================================================

def build_styles(config: Dict[str, Any]) -> StyleSheet1:
    accent = config["branding"]["accent_color"]
    muted = config["branding"]["muted_color"]
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=HexColor(accent),
        alignment=TA_LEFT,
        spaceAfter=10,
    ))
    styles.add(ParagraphStyle(
        name="SectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=HexColor(accent),
        spaceBefore=10,
        spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="BodySmall",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=13,
        textColor=HexColor(config["branding"]["text_color"]),
        spaceAfter=5,
    ))
    styles.add(ParagraphStyle(
        name="Muted",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=HexColor(muted),
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="KPIValue",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        alignment=TA_CENTER,
        textColor=HexColor(accent),
    ))
    styles.add(ParagraphStyle(
        name="KPILabel",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        alignment=TA_CENTER,
        textColor=HexColor(config["branding"]["text_color"]),
    ))
    return styles


def kpi_cards(ctx: ReportContext, styles: StyleSheet1) -> Table:
    currency = ctx.config["pdf"]["currency_symbol"]
    items = [
        ("Receita total", fmt_currency(ctx.kpis.get("receita_total"), currency)),
        ("Lucro total", fmt_currency(ctx.kpis.get("lucro_total"), currency)),
        ("Margem de lucro", fmt_percent(ctx.kpis.get("margem_lucro"))),
        ("Ticket médio", fmt_currency(ctx.kpis.get("ticket_medio"), currency)),
    ]
    cells = []
    for label, value in items:
        cells.append(Paragraph(f"<para align='center'><b>{value}</b><br/>{label}</para>", styles["BodySmall"]))
    table = Table([cells], colWidths=[4.2 * cm] * 4, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HexColor(ctx.config["branding"]["secondary_color"])),
        ("BOX", (0, 0), (-1, -1), 0.7, HexColor(ctx.config["branding"]["accent_color"])),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    return table


def narrative_summary(ctx: ReportContext) -> str:
    receita = fmt_currency(ctx.kpis.get("receita_total"), ctx.config["pdf"]["currency_symbol"])
    lucro = fmt_currency(ctx.kpis.get("lucro_total"), ctx.config["pdf"]["currency_symbol"])
    margem = fmt_percent(ctx.kpis.get("margem_lucro"))
    atraso = fmt_percent(ctx.kpis.get("probabilidade_atraso_pagamento"))
    return (
        f"O negócio <b>{ctx.negocio_nome}</b>, do segmento <b>{ctx.segmento}</b>, apresentou "
        f"receita total de <b>{receita}</b>, lucro total de <b>{lucro}</b> e margem estimada de "
        f"<b>{margem}</b>. A probabilidade empírica de atraso em pagamentos ficou em <b>{atraso}</b>. "
        "O relatório consolida os resultados gerados pelo Insight Calculado e organiza os principais "
        "achados em formato executivo para apoiar análise, acompanhamento e comunicação gerencial."
    )


def lesson_table(ctx: ReportContext, styles: StyleSheet1) -> Table:
    rows = [["Aula", "Tema", "Fixação no sistema financeiro"]]
    for item in ctx.aula_cards:
        rows.append([
            item["aula"],
            Paragraph(item["tema"], styles["BodySmall"]),
            Paragraph(item["insight_principal"] or "-", styles["BodySmall"]),
        ])
    table = Table(rows, colWidths=[2.0 * cm, 5.0 * cm, 9.0 * cm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(ctx.config["branding"]["accent_color"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, HexColor("#D1D5DB")),
        ("BOX", (0, 0), (-1, -1), 0.6, HexColor("#9CA3AF")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return table


def correlation_table(ctx: ReportContext, styles: StyleSheet1) -> Table:
    rows = [["Relação", "Correlação", "Leitura"]]
    for item in build_correlation_rows(ctx):
        val = "-" if item["correlacao"] is None else fmt_number(item["correlacao"], 4)
        rows.append([item["relacao"], val, item["forca"]])
    table = Table(rows, colWidths=[7.2 * cm, 3.0 * cm, 5.0 * cm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(ctx.config["branding"]["accent_color"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, HexColor("#D1D5DB")),
        ("BOX", (0, 0), (-1, -1), 0.6, HexColor("#9CA3AF")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def top_categories_table(ctx: ReportContext, styles: StyleSheet1) -> Table:
    rows = [["Categoria", "Frequência", "Percentual"]]
    for item in build_top_categories_rows(ctx):
        rows.append([item["categoria"], item["frequencia"], f"{fmt_number(item['percentual'], 2)}%"])
    table = Table(rows, colWidths=[7.0 * cm, 4.0 * cm, 4.2 * cm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor(ctx.config["branding"]["accent_color"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, HexColor("#D1D5DB")),
        ("BOX", (0, 0), (-1, -1), 0.6, HexColor("#9CA3AF")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def _pdf_header_footer(canvas, doc, ctx: ReportContext):
    canvas.saveState()
    accent = HexColor(ctx.config["branding"]["accent_color"])
    muted = HexColor(ctx.config["branding"]["muted_color"])
    width, height = A4

    canvas.setStrokeColor(accent)
    canvas.setLineWidth(2)
    canvas.line(doc.leftMargin, height - 1.4 * cm, width - doc.rightMargin, height - 1.4 * cm)

    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(accent)
    canvas.drawString(doc.leftMargin, height - 1.0 * cm, ctx.config["branding"]["company"])

    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(muted)
    canvas.drawRightString(width - doc.rightMargin, 1.0 * cm, f"Página {doc.page}")
    canvas.drawString(doc.leftMargin, 1.0 * cm, ctx.config["branding"]["footer_text"])
    canvas.restoreState()


def export_pdf(ctx: ReportContext, path: Path, chart_dir: Path) -> Path:
    ensure_dir(chart_dir)
    styles = build_styles(ctx.config)

    line_chart = create_line_chart(ctx, chart_dir / "series_financeiras.png")
    cat_chart = create_top_categories_chart(ctx, chart_dir / "top_categorias.png")

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        topMargin=2.2 * cm,
        bottomMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        rightMargin=1.8 * cm,
    )

    story = []

    if ctx.config["pdf"]["include_cover"]:
        story.append(Paragraph(ctx.config["branding"]["report_title"], styles["ReportTitle"]))
        story.append(Paragraph(ctx.config["branding"]["subtitle"], styles["BodySmall"]))
        if ctx.config["pdf"]["show_confidentiality_badge"]:
            story.append(Paragraph(f"<b>Classificação:</b> {ctx.config['branding']['confidentiality']}", styles["Muted"]))
        story.append(Spacer(1, 0.3 * cm))

        meta = [
            ["Negócio", ctx.negocio_nome],
            ["Segmento", ctx.segmento],
            ["Cidade", ctx.cidade],
            ["Gerado em", ctx.created_at],
        ]
        meta_table = Table(meta, colWidths=[3.5 * cm, 11.8 * cm])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), HexColor(ctx.config["branding"]["secondary_color"])),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOX", (0, 0), (-1, -1), 0.5, HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, HexColor("#E5E7EB")),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("Resumo executivo", styles["SectionHeading"]))
        story.append(Paragraph(narrative_summary(ctx), styles["BodySmall"]))
        story.append(Spacer(1, 0.3 * cm))
        story.append(kpi_cards(ctx, styles))
        story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("Indicadores e sinais de atenção", styles["SectionHeading"]))
    if ctx.alertas:
        for alerta in ctx.alertas:
            story.append(Paragraph(f"• {alerta}", styles["BodySmall"]))
    else:
        story.append(Paragraph("Não foram identificados alertas críticos no conjunto de saída utilizado para este exemplo.", styles["BodySmall"]))

    story.append(Spacer(1, 0.25 * cm))
    story.append(Paragraph("Correlações observadas", styles["SectionHeading"]))
    story.append(Paragraph(
        "A tabela abaixo resume as relações estatísticas destacadas pelo Insight Calculado. "
        "Essas associações ajudam a selecionar variáveis com maior potencial explicativo, mas não substituem validação contextual.",
        styles["BodySmall"]
    ))
    story.append(correlation_table(ctx, styles))
    story.append(Spacer(1, 0.4 * cm))

    if ctx.config["pdf"]["include_charts"]:
        story.append(Paragraph("Visualizações financeiras", styles["SectionHeading"]))
        story.append(Paragraph("Os gráficos sintetizam a evolução dos valores por período e a concentração por categoria observada.", styles["BodySmall"]))
        story.append(Image(str(line_chart), width=16.2 * cm, height=7.2 * cm))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Image(str(cat_chart), width=15.5 * cm, height=7.0 * cm))
        story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("Categorias de maior recorrência", styles["SectionHeading"]))
    story.append(top_categories_table(ctx, styles))

    if ctx.config["pdf"]["include_lesson_breakdown"]:
        story.append(PageBreak())
        story.append(Paragraph("Construção pedagógica do sistema financeiro", styles["SectionHeading"]))
        story.append(Paragraph(
            "Cada aula consolida uma camada do sistema financeiro: coleta, estatística descritiva, distribuição, correlação, "
            "probabilidade, comparação de modelos, tratamento de ruídos e visualização final.",
            styles["BodySmall"]
        ))
        story.append(lesson_table(ctx, styles))

    if ctx.config["pdf"]["include_methodology"]:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph("Observações metodológicas", styles["SectionHeading"]))
        methodology = [
            "Os valores apresentados são lidos diretamente do JSON de saída do Insight Calculado.",
            "Os arquivos CSV priorizam intercâmbio de dados; o Excel organiza a informação em múltiplas abas; o PDF prioriza comunicação executiva.",
            "O relatório pode ser personalizado por meio de um arquivo JSON de configuração, com ajustes de identidade visual, textos e seções exibidas.",
            safe_get(ctx.raw, "aula_8", "calculos", "normas_e_legislacao_nota", default="Os resultados devem observar normas e legislações aplicáveis."),
        ]
        for item in methodology:
            story.append(Paragraph(f"• {item}", styles["BodySmall"]))

    doc.build(story, onFirstPage=lambda c, d: _pdf_header_footer(c, d, ctx), onLaterPages=lambda c, d: _pdf_header_footer(c, d, ctx))
    return path


# ============================================================
# Orquestração
# ============================================================

def generate_reports(input_json: Path, output_dir: Path, config_json: Optional[Path] = None, prefix: Optional[str] = None) -> Dict[str, str]:
    data = load_json(input_json)
    config = DEFAULT_CONFIG
    if config_json and config_json.exists():
        user_cfg = load_json(config_json)
        config = deep_merge(DEFAULT_CONFIG, user_cfg)

    ctx = build_context(data, config)
    output_dir = output_dir.resolve()
    ensure_dir(output_dir)
    prefix = prefix or slugify(ctx.negocio_nome)

    csv_dir = output_dir / "csv"
    charts_dir = output_dir / "_charts"
    ensure_dir(csv_dir)
    ensure_dir(charts_dir)

    csv_files = export_csv_bundle(ctx, csv_dir, prefix)
    excel_file = export_excel(ctx, output_dir / f"{prefix}_relatorio_financeiro.xlsx")
    pdf_file = export_pdf(ctx, output_dir / f"{prefix}_relatorio_financeiro.pdf", charts_dir)

    manifest = {
        "input_json": str(input_json),
        "output_dir": str(output_dir),
        "csv_files": [str(p) for p in csv_files],
        "excel_file": str(excel_file),
        "pdf_file": str(pdf_file),
        "config_used": config,
        "generated_at": ctx.created_at,
    }
    save_json(output_dir / f"{prefix}_manifesto_relatorios.json", manifest)
    return manifest


def prompt_for_path(prompt_text: str, must_exist: bool = False, expect_file: bool = False) -> Path:
    while True:
        raw = input(prompt_text).strip().strip(""").strip("'")
        if not raw:
            print("Entrada vazia. Tente novamente.")
            continue
        path = Path(raw).expanduser()
        if must_exist and not path.exists():
            print(f"Caminho não encontrado: {path}")
            continue
        if expect_file and path.exists() and not path.is_file():
            print(f"O caminho informado não é um arquivo: {path}")
            continue
        return path


def prompt_optional_path(prompt_text: str) -> Optional[Path]:
    raw = input(prompt_text).strip().strip(""").strip("'")
    if not raw:
        return None
    path = Path(raw).expanduser()
    if not path.exists():
        print(f"Aviso: o arquivo de configuração não foi encontrado e será ignorado: {path}")
        return None
    return path


def _create_tk_root() -> Optional[Any]:
    if tk is None:
        return None
    try:
        root = tk.Tk()
        root.withdraw()
        try:
            root.attributes("-topmost", True)
        except Exception:
            pass
        return root
    except Exception:
        return None


def choose_paths_with_windows_dialog() -> Tuple[Path, Path, Optional[Path]]:
    root = _create_tk_root()
    if root is None or filedialog is None:
        raise RuntimeError("Interface gráfica indisponível para seleção de arquivos.")

    try:
        input_file = filedialog.askopenfilename(
            title="Selecione o JSON de saída do Insight Calculado",
            filetypes=[("Arquivos JSON", "*.json"), ("Todos os arquivos", "*.*")],
        )
        if not input_file:
            raise SystemExit("Seleção cancelada pelo usuário.")

        output_dir = filedialog.askdirectory(
            title="Selecione a pasta onde os relatórios serão gerados"
        )
        if not output_dir:
            raise SystemExit("Seleção cancelada pelo usuário.")

        config_file = filedialog.askopenfilename(
            title="Selecione o JSON de configuração visual (opcional) ou clique em Cancelar",
            filetypes=[("Arquivos JSON", "*.json"), ("Todos os arquivos", "*.*")],
        )

        return (
            Path(input_file).expanduser(),
            Path(output_dir).expanduser(),
            Path(config_file).expanduser() if config_file else None,
        )
    finally:
        try:
            root.destroy()
        except Exception:
            pass


def main():
    if len(sys.argv) >= 3:
        input_json = Path(sys.argv[1]).expanduser()
        output_dir = Path(sys.argv[2]).expanduser()
        config_json = Path(sys.argv[3]).expanduser() if len(sys.argv) >= 4 else None
        prefix = sys.argv[4] if len(sys.argv) >= 5 else None
    else:
        print("Gerador de relatórios do Insight Calculado")
        print("Uma janela do Explorador de Arquivos será aberta para selecionar o JSON de entrada e a pasta de saída.\n")
        try:
            input_json, output_dir, config_json = choose_paths_with_windows_dialog()
        except Exception as exc:
            print(f"Não foi possível abrir o seletor gráfico ({exc}).")
            print("Será usado o modo de digitação manual.\n")
            input_json = prompt_for_path(
                "Digite o caminho do arquivo JSON de saída do Insight Calculado: ",
                must_exist=True,
                expect_file=True,
            )
            output_dir = prompt_for_path(
                "Digite a pasta onde os relatórios serão gerados: ",
                must_exist=False,
                expect_file=False,
            )
            config_json = prompt_optional_path(
                "Digite o caminho do JSON de configuração visual (ou pressione Enter para usar o padrão): "
            )
        prefix = input(
            "Digite um prefixo para os nomes dos arquivos (ou pressione Enter para usar o nome do negócio): "
        ).strip() or None

    manifest = generate_reports(input_json, output_dir, config_json, prefix)
    print("\nRelatórios gerados com sucesso:")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
